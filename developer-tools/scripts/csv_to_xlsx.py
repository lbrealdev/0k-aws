#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "pandas",
#   "openpyxl",
# ]
# ///

import os
import sys
import pandas as pd
from openpyxl import Workbook


def main():
    if len(sys.argv) != 2:
        print("Usage: uv run csv_to_xlsx.py <report_directory>")
        sys.exit(1)

    csv_dir = sys.argv[1]
    if not os.path.isdir(csv_dir):
        print(f"Error: Directory '{csv_dir}' not found.")
        sys.exit(1)

    xlsx_file = f"{csv_dir}.xlsx"

    sheet_names = {
        "aws-cc": "CodeCommit",
        "aws-ca": "CodeArtifact",
        "aws-cb": "CodeBuild",
        "aws-cd": "CodeDeploy",
        "aws-cp": "CodePipeline",
    }

    wb = Workbook()
    wb.remove(wb.active)

    csv_count = 0
    for file in os.listdir(csv_dir):
        if file.endswith(".csv"):
            prefix = file.split("-")[1]
            sheet_name = sheet_names.get(f"aws-{prefix}", f"Unknown_{prefix}")

            df = pd.read_csv(os.path.join(csv_dir, file))
            if df.empty:
                continue

            ws = wb.create_sheet(title=sheet_name)
            for r, (idx, row) in enumerate(df.iterrows(), start=2):
                for c, value in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=value)
            for c, col in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c, value=col)

            csv_count += 1

    if csv_count > 0:
        wb.save(xlsx_file)
        print(f"XLSX created: {xlsx_file} ({csv_count} worksheets)")
    else:
        print("No valid CSV files found to convert.")


if __name__ == "__main__":
    main()
